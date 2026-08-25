"""Generate the Fall 2026 PHY 210 (Mathematical Physics, Smith) course calendar.

Format follows EarlhamArtifacts/P125 F2025 Calendar.xlsx: a "Schedule" sheet
with Week | Class | Date | Topics | Reading Due | HW Due | Exams, in two
side-by-side half-semester blocks, plus a "Grade Categories" sheet skeleton.
Content follows the previous Smith professor's Spring '26 PHY 210 sequence
(Felder & Felder), remapped onto the Smith Fall 2026 academic calendar.

Usage: python make_fall2026_calendar.py OUTPUT.xlsx
"""
import sys
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------- semester
# Smith Fall 2026: classes Tue Sep 8 - Tue Dec 15.
# MWF meetings; skip Mon Oct 12 + Tue Oct 13 (autumn recess),
# Wed Nov 25 + Fri Nov 27 (Thanksgiving). Cromwell Day (Tue Nov 10) and
# the Dec 15 last-day-of-classes Tuesday don't hit a MWF pattern.
FIRST_DAY = date(2026, 9, 9)   # first MWF meeting (classes open Tue Sep 8)
LAST_DAY = date(2026, 12, 14)  # last MWF meeting
NO_CLASS = {
    date(2026, 10, 12): "Autumn recess",
    date(2026, 11, 25): "Thanksgiving",
    date(2026, 11, 27): "Thanksgiving",
}

def class_days():
    d = FIRST_DAY
    while d <= LAST_DAY:
        if d.weekday() in (0, 2, 4) and d not in NO_CLASS:  # M W F
            yield d
        d += timedelta(days=1)

# ------------------------------------------------------------------ content
# (topic, reading_due) per teaching slot, in the previous prof's order.
# Readings are the Felder & Felder sections being covered; the "Reading Due"
# column mirrors the P125 calendar's read-before-class convention.
#
# Quizzes are dedicated in-class days (as on the previous prof's calendar),
# pinned to Friday slot indices; the flex day sits where her Spring '26
# snow day fell (week 5) and absorbs snow / Mountain Day / drift.
# 35 content slots + 3 quiz days + 1 flex day = 39 MWF meetings, exactly
# matching her 39 spring slots. The calendar is FULL: the flex day is the
# only buffer left (the old practice/review day is now the coordinate-
# systems day), so adding anything else means displacing content.
SPECIALS = {
    7: ("Quiz 1 (Ch 1)", True),                          # Fri Sep 25
    13: ("Flex day (snow / Mountain Day buffer)", False),  # Fri Oct 9
    18: ("Quiz 2 (Ch 10, 3, 2)", True),                  # Fri Oct 23
    30: ("Quiz 3 (Ch 6, 5)", True),                      # Fri Nov 20
}
CONTENT = [
    ("Syllabus; SHO and overview of differential equations", "1.1-1.2"),
    ("Generating ODEs from physical situations", "1.1-1.2"),
    ("Arbitrary constants; initial conditions", "1.3"),
    ("Separation of variables", "1.5"),
    ("Guess and check", "1.6"),
    ("Linearity, homogeneity, superposition", "1.6"),
    ("Methods of solving ODEs: guess and check", "10.2"),
    ("Jupyter notebook exercise: ODEs in Python", "10.2"),
    ("Heaviside, Dirac delta, and the Laplace transform", "10.10"),
    ("Using Laplace transforms to solve ODEs", "10.11"),
    ("Complex numbers: basic properties", "3.1-3.4"),
    ("Euler's formula; complex ODEs", "3.5"),
    ("Linear approximations", "2.1-2.2"),
    ("Maclaurin series", "2.3"),
    # 2.6-2.7 (convergence) deliberately cut (decision 2026-08-17);
    # Appendix C is the pointed-to substitute.
    ("Taylor series; finding one series from another", "2.4-2.5 (convergence: App. C)"),
    ("Properties of matrices", "6.1-6.2"),
    ("Matrix x column; vector transformations", "6.3-6.4"),
    ("Matrix multiplication; identity, determinant, inverse", "6.5-6.7"),
    ("Finding eigenvalues and eigenvectors", "6.8"),
    ("Eigenvalues and eigenvectors, continued", "6.8"),
    ("The two-coupled-oscillator problem", "6.9"),
    ("Setting up 1D and 2D integrals", "5.1-5.2"),
    # Polar double integrals are Felder 5.6, not 5.3-5.4 (verified against the
    # textbook TOC 2026-08-17; the S26 grid this row was copied from mislabels it).
    ("Cartesian 2D integrals; polar coordinates", "5.3-5.4, 5.6"),
    ("Line integrals and surface integrals", "5.8, 5.10"),
    # Decision 2026-08-17: the review day IS the coordinate-systems day
    # (how the S26 sections actually used it); 5.5/5.7 get real coverage
    # and Quiz 3 can fairly test them.
    ("Coordinate systems: cylindrical and spherical (review + practice)",
     "5.5, 5.7; App. D"),
    ("Vector and scalar fields; the gradient", "8.1-8.4"),
    ("Work, path integrals, and the gradient theorem", "8.5"),
    # Decision 2026-08-17: Feynman's geometric definitions of divergence
    # and curl come BEFORE Felder's treatment. Free reading edition:
    # feynmanlectures.caltech.edu (Vol II Ch 2-3).
    ("Divergence and curl, geometrically", "Feynman Vol II Ch 2-3"),
    ("Divergence, curl, and the Laplacian", "8.6-8.7"),
    ("Divergence theorem; Stokes' theorem", "8.9-8.10; Feynman Vol II Ch 3"),
    ("Conservative vector fields", "8.11"),
    # Decisions 2026-08-17: Fourier series in 2 days; PDEs protected with
    # 2 days; Fourier transforms ride the finale (graphical treatment +
    # exoplanets / image compression as the closing demos, a la Gary
    # Felder's last-day design).
    ("Introduction to Fourier series", "9.1-9.3"),
    ("Fourier series: different periods, finite domains, complex exponentials",
     "9.4-9.5"),
    ("Intro to PDEs: the heat equation; separation of variables",
     "11.1-11.2, 11.4"),
    ("Normal modes of the wave equation; Fourier transforms",
     "11.3, 9.6"),
]

# Extra (non-WHW) due dates shown in the HW Due column.
# Non-Newtonian Scientist: mid-semester (decision 2026-08-17); Mon Oct 26
# is the calendar midpoint and has no competing WHW deadline.
EXTRA_DUE = {
    date(2026, 10, 26): "Non-Newtonian Scientist due",
}

# ---------------------------------------------------------------- PCCIs
# Pre-Class Check-Ins, Gary-Felder style (decision 2026-08-17): due at
# the start of class most days; mostly the section's Discovery Exercise
# (DE x.y.1 = the Discovery Exercise of section x.y), done BEFORE the
# topic. Quiz days and the flex day get none. Every problem/DE number
# below is attested in a predecessor's actual assignment (Gary's PCCI
# schedules, Gillian's WHW forms, Casey's in-class problem sets) -- see
# private/GillianManbirS26/.../WHW problem lists (transcribed).md and
# the 2026-08-17 inventory report. Timing calibration: Gary's students
# reported ~8-27 min per DE (old/15F/Assigned Problems.docx).
PCCI = {
    # No PCCI on day 1 (nobody has the syllabus before the first
    # class); first collection is day 2, matching Gary's practice.
    # DE 1.2.1 is deliberately unassigned: class 01 does its content
    # live (the y'=8x ladder).
    date(2026, 9, 11): "Read the syllabus and the advice from former "
                       "students; bring one question or comment",
    date(2026, 9, 14): "DE 1.3.1 Part 1",
    date(2026, 9, 16): "DE 1.5.1 Parts 1-5",
    date(2026, 9, 18): "DE 1.6.1 Parts 1-3",
    date(2026, 9, 21): "1.166",
    date(2026, 9, 23): "10.1, 10.3",
    date(2026, 9, 28): "Log into posit.smith.edu and open a blank "
                       "Jupyter notebook; bring your laptop",
    date(2026, 9, 30): "10.5, 10.8",
    date(2026, 10, 2): "10.216",
    date(2026, 10, 5): "DE 3.2.1",
    date(2026, 10, 7): "DE 3.4.1",
    date(2026, 10, 14): "DE 2.2.1 Parts 1-5",
    date(2026, 10, 16): "DE 2.3.1 Parts 1-3",
    date(2026, 10, 19): "2.209",
    date(2026, 10, 21): "6.2",
    date(2026, 10, 26): "DE 6.3.1",
    date(2026, 10, 28): "DE 6.5.1",
    date(2026, 10, 30): "DE 6.8.1 Parts 1-3",
    date(2026, 11, 2): "6.166",
    date(2026, 11, 4): "6.175",
    date(2026, 11, 6): "5.1",
    date(2026, 11, 9): "DE 5.4.1",
    date(2026, 11, 11): "5.163 parts a and d",
    date(2026, 11, 13): "DE 5.7.1 Parts 1-4",
    date(2026, 11, 16): "DE 8.2.1 Parts 5-6",
    date(2026, 11, 18): "DE 8.4.1",
    date(2026, 11, 23): "Read Feynman II Ch 2; write one sentence on "
                        "what the gradient of temperature means "
                        "physically",
    date(2026, 11, 30): "DE 8.6.1 Parts 1-9",
    date(2026, 12, 2): "DE 8.9.1",
    date(2026, 12, 4): "DE 8.11.1 Parts 1-2, 4-5, 7-8",
    date(2026, 12, 7): "DE 9.2.1 Parts 1-3",
    date(2026, 12, 9): "DE 9.4.1 Part 1",
    date(2026, 12, 11): "DE 11.2.1 Parts 1-2",
    date(2026, 12, 14): "DE 11.3.1 Part 1",
}

# ------------------------------------------------------- WHW problem lists
# 13 weekly lists, re-cut from Gillian's S26 WHW forms (transcribed
# 2026-08-17) onto the F2026 week boundaries: WHW N covers the Monday +
# Wednesday material of its due week, with Friday's new topic rolling
# into the next WHW (the S26 convention). Where our calendar covers
# material S26's forms did not (coordinates day, the Feynman div/curl
# days, 9.6, 11.3), problems come from Gary Felder's assigned lists and
# Casey Berger's F22 in-class sets -- every number is attested.
# Problems assigned as a PCCI are NOT repeated in the WHW lists.
# Tuples: (whw, covers, warmup, essentials, depth).
WHWS = [
    (1, "Intro to ODEs (1.1-1.2)",
     "ODEs: 1.17, 1.19",
     "ODEs: 1.18, 1.20, 1.21, 1.25",
     "ODEs: 1.33, 1.36"),
    (2, "Arbitrary constants (1.3); separation of variables (1.5)",
     "Arbitrary constants: 1.39, 1.41, 1.43. "
     "Separation of variables: 1.90, 1.91, 1.93, 1.95",
     "Arbitrary constants: 1.38, 1.46, 1.58. "
     "Separation of variables: 1.103",
     "Arbitrary constants: 1.60. Separation of variables: 1.105"),
    (3, "Guess and check, superposition (1.6); linear operators (10.2)",
     "Guess and check: 1.108, 1.109",
     # 1.129 (linearity/homogeneity/superposition tester), NOT 1.29 (a
     # slope-fields problem from cut section 1.4) -- transcription fix
     # 2026-08-24 from the S26 form screenshots.
     "Guess and check: 1.111, 1.113, 1.117, 1.122, 1.129. "
     "Sec 10.2: 10.13, 10.26, 10.31",
     "Guess and check: 1.125, 1.131. Sec 10.2: 10.32, 10.33"),
    (4, "ODEs in Python; Heaviside, Dirac, Laplace (10.10)",
     "10.217, 10.218",
     "Sec 10.10: 10.219, 10.223, 10.230, 10.242",
     "Redo the class notebook's exercises from scratch in a fresh "
     "notebook on posit.smith.edu"),
    (5, "Solving ODEs with Laplace transforms (10.11); "
        "complex numbers and Euler (3.1-3.5)",
     "Sec 10.11: 10.246, 10.248. Complex numbers: 3.17, 3.19, 3.47. "
     "Euler / complex ODE: 3.59, 3.65, 3.92",
     "Sec 10.11: 10.252, 10.261, 10.263, 10.264. "
     "Complex numbers: 3.49, 3.54. Euler / complex ODE: 3.77, 3.94, 3.95",
     "Sec 10.11: 10.270, 10.272, 10.274. Complex numbers: 3.45, 3.56. "
     "Euler / complex ODE: 3.85, 3.107"),
    (6, "Linear approximations (2.1-2.2); Maclaurin series (2.3)",
     "Linear approximations: 2.6, 2.7, 2.16",
     "Maclaurin series: 2.31, 2.35, 2.37",
     "Maclaurin series: 2.86, 2.87"),
    (7, "Taylor series (2.4-2.5); matrices and the three-spring "
        "problem (6.1-6.2)",
     "Taylor series: 2.41. Normal modes: 6.3. Matrices: 6.19",
     "Maclaurin/Taylor: 2.47, 2.49, 2.53, 2.55, 2.57, 2.67. "
     "The three-spring problem: 6.7, 6.17. Normal modes: 6.9. "
     "Matrices: 6.21, 6.23",
     "Taylor series: 2.71, 2.94"),
    (8, "Matrix times column, basis, matrix times matrix, identity, "
        "inverse, determinants (6.3-6.7)",
     "Matrix times column: 6.27, 6.29, 6.31. Matrix-matrix: 6.69. "
     "Identity and inverse: DE 6.6.1, 6.97. Determinants: 6.127, 6.129",
     "Matrix times column: 6.35, 6.41. Matrix-matrix: 6.81, 6.89, 6.91. "
     "Identity and inverse: 6.99, 6.101, 6.103, 6.107, 6.109, 6.119, "
     "6.121, 6.123. Determinants: 6.133, 6.139",
     "Basis and transformations: 6.45, 6.53, 6.55, 6.57, 6.61. "
     "Determinants: 6.147, 6.161"),
    (9, "Eigenvalues and eigenvectors (6.8); coupled oscillators (6.9)",
     "Eigenvectors & eigenvalues: 6.170, 6.172",
     "Eigenvectors & eigenvalues: 6.171, 6.173, 6.177",
     "Eigenvectors & eigenvalues: 6.179, 6.181, 6.185"),
    (10, "Setting up integrals; Cartesian doubles, polar; line and "
         "surface integrals (5.1-5.4, 5.6, 5.8, 5.10)",
     "Setting up 1D integrals: 5.2, 5.3. Single integrals in multiple "
     "dimensions: 5.25, 5.27, 5.29, 5.31. Cartesian rectangular double "
     "integrals: 5.61, 5.63. Line integrals: 5.197. Surface integrals: "
     "5.255",
     "Setting up 1D integrals: 5.7, 5.9, 5.11, 5.19, 5.23. Single "
     "integrals in multiple dimensions: 5.35, 5.43, 5.45, 5.51. "
     "Cartesian rectangular: 5.71, 5.75, 5.77. Cartesian "
     "non-rectangular: 5.83, 5.85, 5.97, 5.101. Line integrals: 5.199, "
     "5.201, 5.215, 5.217, 5.229. Surface integrals: 5.256, 5.257, 5.263",
     "Line integrals: 5.223, 5.225, 5.235"),
    (11, "Cylindrical and spherical coordinates (5.5, 5.7); fields, "
         "potential, gradient (8.1-8.5)",
     "Spherical coordinates: 5.167, 5.174. Scalar and vector fields: "
     "8.1. Potential in 1D: DE 8.3.1, 8.29, 8.35. From "
     "potential to gradients: 8.53, 8.55",
     "Spherical coordinates: 5.181, 5.187. Polar and cylindrical: "
     "5.129, 5.137, 5.147, 5.153. Scalar and vector fields: 8.5, 8.7, "
     "8.9, 8.17, 8.19, 8.21, 8.25. Potential in 1D: 8.37, 8.41, 8.43. "
     "From potential to gradients: 8.57, 8.59, 8.65",
     "Polar and cylindrical: 5.157, 5.169, 5.171, 5.193. "
     "All coordinates: 5.165"),
    (12, "Divergence and curl, Feynman and Felder (Feynman II 2-3, "
         "8.6-8.7); divergence theorem and Stokes' theorem (8.9-8.10)",
     "Divergence and curl by inspection: 8.84, 8.86",
     "Divergence and curl: 8.88, 8.90, 8.94, 8.98, 8.102, 8.110. "
     "Divergence theorem: 8.147, 8.151, 8.155, 8.157. "
     "Stokes' theorem: 8.159, 8.163, 8.165, 8.169, 8.171",
     "8.106. Write out Feynman's flux-through-a-tiny-cube derivation "
     "of the divergence theorem in your own words, with pictures"),
    (13, "Conservative fields (8.11); Fourier series (9.1-9.5)",
     "Fourier series: 9.5, 9.7, 9.9",
     "Conservative fields: 8.178. Fourier series: 9.15, 9.17, 9.23, "
     "9.27, 9.29, 9.33",
     "Conservative fields: 8.173, 8.177, 8.179, 8.181. Different "
     "periods and finite domains: 9.38, 9.41, 9.47. Complex Fourier "
     "series: 9.60, 9.62, 9.66"),
]

def build(outpath):
    days = list(class_days())
    n = len(days)
    assert len(CONTENT) + len(SPECIALS) == n, (
        f"{len(CONTENT)} content + {len(SPECIALS)} specials "
        f"for {n} class meetings")
    assert all(days[i].weekday() == 4 for i in SPECIALS), (
        "quiz/flex day not on a Friday")
    special_days = {days[i] for i in SPECIALS}
    assert all(d in days for d in PCCI), "PCCI assigned to a non-class day"
    assert not any(d in PCCI for i, d in enumerate(days)
                   if i in SPECIALS and SPECIALS[i][1]), (
        "PCCI assigned to a quiz day")

    # rows: one per class meeting; insert break markers
    rows = []      # (week, class_no, date, topic, reading, pcci, hw, exam)
    week_no = 0
    last_week = None
    class_no = 0
    hw_no = 0
    content_i = 0
    breaks_seen = set()
    whw_due = {}   # WHW number -> due date, for the WHW sheet
    for slot_i, d in enumerate(days):
        iso_week = d.isocalendar()[1]
        if iso_week != last_week:
            week_no += 1
            last_week = iso_week
        # note upcoming breaks as their own marker rows
        for bd, why in NO_CLASS.items():
            if bd not in breaks_seen and bd < d:
                breaks_seen.add(bd)
                rows.append((None, None, bd, f"No class - {why}",
                             "", "", "", ""))
        if slot_i in SPECIALS:
            label, is_quiz = SPECIALS[slot_i]
            topic, reading, exam = label, "", (label if is_quiz else "")
        else:
            topic, reading = CONTENT[content_i]
            exam = ""
            content_i += 1
        hw = ""
        if d.weekday() == 4 and slot_i > 0:  # Fridays (incl. quiz days,
            hw_no += 1                       # matching the previous prof)
            hw = f"WHW{hw_no:02d}"
            whw_due[hw_no] = d
        if d in EXTRA_DUE:
            hw = f"{hw}; {EXTRA_DUE[d]}" if hw else EXTRA_DUE[d]
        class_no += 1
        rows.append((week_no, class_no, d, topic, reading,
                     PCCI.get(d, ""), hw, exam))
    for bd, why in NO_CLASS.items():
        if bd not in breaks_seen:
            rows.append((None, None, bd, f"No class - {why}",
                         "", "", "", ""))
    rows.sort(key=lambda r: r[2])
    rows.append((None, None, date(2026, 12, 19),
                 "Final exam period Dec 19-22 (registrar schedules)",
                 "", "", "", "Final (Ch 8, 9, 11 + redemptions)"))
    assert set(whw_due) == {w[0] for w in WHWS}, (
        f"WHW sheet numbers {sorted(w[0] for w in WHWS)} != "
        f"calendar WHW numbers {sorted(whw_due)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    break_fill = PatternFill("solid", fgColor="FCE4D6")
    exam_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["Week", "Class", "Date", "Topics", "Reading Due", "PCCI",
               "HW Due", "Exams"]
    # two side-by-side blocks, like the P125 calendar: cols A-H and J-Q
    half = (len(rows) + 1) // 2
    blocks = [rows[:half], rows[half:]]
    for b, block in enumerate(blocks):
        c0 = 1 + b * 9  # A=1, J=10
        for j, h in enumerate(headers):
            cell = ws.cell(row=1, column=c0 + j, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        for i, (wk, cn, d, topic, reading, pcci, hw, exam) in enumerate(
                block, start=2):
            vals = [wk, cn, d.strftime("%a %b %-d"), topic, reading,
                    pcci, hw, exam]
            for j, v in enumerate(vals):
                cell = ws.cell(row=i, column=c0 + j, value=v)
                cell.border = border
                cell.alignment = wrap
                if cn is None:
                    cell.fill = break_fill
                elif exam:
                    cell.fill = exam_fill

    widths = [6, 6, 11, 40, 12, 20, 12, 20]
    for b in range(2):
        for j, w in enumerate(widths):
            col = openpyxl.utils.get_column_letter(1 + b * 9 + j)
            ws.column_dimensions[col].width = w

    # ----------------------------------------------- WHW problem lists
    whw = wb.create_sheet("WHW Problem Lists")
    note = ("The following problems are useful practice for the week. "
            "I recommend starting with the warm-ups. If they're too "
            "easy, move on to Essentials. If you're interested in the "
            "Depth content, you can add those problems as well. You "
            "are not expected to do all the problems. DE x.y.1 = the "
            "Discovery Exercise of section x.y. Answers to odd "
            "problems: felderbooks.com/mathmethods (Appendix M).")
    whw.append([note])
    whw.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    whw.cell(row=1, column=1).alignment = wrap
    whw.row_dimensions[1].height = 60
    whw.append(["WHW", "Due", "Covers", "Warm-up", "Essentials", "Depth"])
    for cell in whw[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for i, (wn, covers, warm, ess, depth) in enumerate(WHWS, start=3):
        vals = [f"WHW{wn:02d}", whw_due[wn].strftime("%a %b %-d"),
                covers, warm, ess, depth]
        for j, v in enumerate(vals, start=1):
            cell = whw.cell(row=i, column=j, value=v)
            cell.border = border
            cell.alignment = wrap
    for j, w in enumerate([8, 11, 30, 34, 44, 34]):
        whw.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w

    gc = wb.create_sheet("Grade Categories")
    gc.append(["Category", "Number", "Drop", "Points Each", "Total Points"])
    for cell in gc[1]:
        cell.font = header_font
        cell.fill = header_fill
    # Michael's decided scheme (2026-08-11; participation weight raised
    # 2026-08-25: 2 pts/day, quizzes 150->140, final 190->185). Course
    # total must be exactly 1000 points. pts_cell, when set, is the
    # formula written to the sheet (Non-Newtonian Scientist is worth one
    # homework, by reference).
    cats = [
        ("Attendance/participation", 39, 4, 2, None),
        ("Written Homework (WHW)", 13, 1, 25, None),
        ("Non-Newtonian Scientist", 1, 0, 25, "=D3"),
        ("Quizzes", 3, 0, 140, None),
        ("Final exam", 1, 0, 185, None),
    ]
    total = sum((num - drop) * pts for _, num, drop, pts, _ in cats)
    assert total == 1000, f"grade categories sum to {total}, not 1000"
    for i, (name, num, drop, pts, pts_cell) in enumerate(cats, start=2):
        gc.append([name, num, drop, pts_cell or pts, f"=(B{i}-C{i})*D{i}"])
    gc.append(["Total", None, None, None, f"=SUM(E2:E{1 + len(cats)})"])
    for j, w in enumerate([28, 9, 7, 12, 13]):
        gc.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w

    wb.save(outpath)
    print(f"wrote {outpath}: {len(rows)} schedule rows, {n} class meetings")

if __name__ == "__main__":
    build(sys.argv[1])
